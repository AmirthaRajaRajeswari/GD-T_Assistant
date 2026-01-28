import json
import os
import sys
import time
import pandas as pd
from PIL import Image
from google import genai
from google.genai import types
from google.genai.errors import ServerError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path

OUTPUT_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("output")
BASE_DIR = Path(__file__).parent

BLOCKS_JSON = OUTPUT_DIR / "blocks.json"
BLOCKS_DIR = OUTPUT_DIR
OUTPUT_EXCEL = OUTPUT_DIR / "gdt_compliance_report.xlsx"
SUMMARY_JSON = OUTPUT_DIR / "summary.json"
RULES_JSON = BASE_DIR / "rules" / "checklist.json"

API_KEY = "AIzaSyDnTUiXINv7hcYoc2pkgTcD1WNVmAMZ28A"
MODEL_NAME = "gemini-2.5-flash"
client = genai.Client(api_key=API_KEY)

with open(BLOCKS_JSON, "r") as f:
    blocks = json.load(f)

with open(RULES_JSON, "r") as f:
    rules = json.load(f)["rules"]

images = []
image_labels = []

for b in blocks:
    img_path = BLOCKS_DIR / f"{b['id']}.png"
    if img_path.exists():
        images.append(Image.open(img_path))
        image_labels.append({
            "block_id": b["id"],
            "block_type": b["type"],
            "view_type": b.get("view_type", "")
        })

prompt = f"""
You are a GD&T compliance checker following ASME Y14.5-2018.

You are given:
- Multiple cropped blocks from ONE engineering drawing
- A checklist of GD&T rules

Block metadata:
{json.dumps(image_labels, indent=2)}

Checklist rules:
{json.dumps(rules, indent=2)}

Task:
For EACH rule:
- Decide YES / NO / NOT_APPLICABLE
- Give a short technical reason
- Give a recommendation to fix the issue if result is NO
- Base decision ONLY on visible content
- Do NOT assume missing info
- Do NOT infer intent

Return STRICT JSON ONLY in this format:
[
  {{
    "rule_id": "R1",
    "result": "YES | NO | NOT_APPLICABLE",
    "reason": "...",
    "recommendation": "..."
  }}
]
"""

max_retries = 3
delay = 10

for attempt in range(max_retries):
    try:
        response = client.models.generate_content(
            model=MODEL_NAME,
            contents=[prompt, *images],
            config=types.GenerateContentConfig(temperature=0)
        )
        break
    except ServerError:
        if attempt == max_retries - 1:
            raise
        print(f"Gemini overloaded. Retrying in {delay}s...")
        time.sleep(delay)
        delay *= 2

try:
    raw = response.text.strip()
    if raw.startswith("```"):
        raw = raw.replace("```json", "").replace("```", "").strip()
    parsed = json.loads(raw)
except Exception:
    raise RuntimeError("Gemini output not valid JSON:\n" + response.text)

def get_severity(rule_meta: dict, result: str) -> str:
    if result != "NO":
        return ""
    category = rule_meta.get("category", "")
    if category in ["DATUM", "FCF", "TOLERANCE"]:
        return "CRITICAL"
    if category in ["GENERAL_GDT", "DIMENSIONING"]:
        return "MAJOR"
    return "MINOR"

rule_map = {r["id"]: r for r in rules}
results = []

for r in parsed:
    meta = rule_map.get(r["rule_id"], {})
    result_val = r.get("result")
    results.append({
        "Rule ID": r["rule_id"],
        "Rule Description": meta.get("description"),
        "Result": result_val,
        "Severity": get_severity(meta, result_val),
        "Reason": r.get("reason"),
        "Recommendation": r.get("recommendation")
    })

df = pd.DataFrame(results)

df.to_excel(OUTPUT_EXCEL, index=False, sheet_name="Compliance Report")
wb = load_workbook(OUTPUT_EXCEL)
ws = wb["Compliance Report"]

green = PatternFill("solid", "C6EFCE")
red = PatternFill("solid", "FFC7CE")
grey = PatternFill("solid", "E7E6E6")
orange = PatternFill("solid", "FFD966")
bold = Font(bold=True)

headers = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
result_col = headers["Result"]
severity_col = headers.get("Severity")

for row in range(2, ws.max_row + 1):
    r = ws.cell(row, result_col)

    r.fill = green if r.value == "YES" else red if r.value == "NO" else grey

    if severity_col:
        s = ws.cell(row, severity_col)
        s.fill = red if s.value == "CRITICAL" else orange if s.value == "MAJOR" else grey

for cell in ws[1]:
    cell.font = bold

total = len(df)
passed = (df["Result"] == "YES").sum()
failed = (df["Result"] == "NO").sum()
na = (df["Result"] == "NOT_APPLICABLE").sum()
applicable = passed + failed
critical = (df["Severity"] == "CRITICAL").sum()
major = (df["Severity"] == "MAJOR").sum()
compliance = round((passed / applicable) * 100, 1) if applicable else 0
risk = "HIGH" if critical > 0 else "MEDIUM" if major > 0 else "LOW"
issues = [
    {
        "rule_id": row["Rule ID"],
        "reason": row["Reason"],
        "recommendation": row["Recommendation"]
    }
    for _, row in df[df["Result"] == "NO"].iterrows()
]

summary_json = {
    "total_rules": int(total),
    "applicable_rules": int(applicable),
    "passed": int(passed),
    "failed": int(failed),
    "not_applicable": int(na),
    "compliance_percent": float(compliance),
    "critical_issues": int(critical),
    "major_issues": int(major),
    "overall_risk": str(risk),
    "issues": issues
}

summary_path = OUTPUT_DIR / "summary.json"

with open(summary_path, "w") as f:
    json.dump(summary_json, f, indent=2)

wb.save(OUTPUT_EXCEL)

print("GD&T compliance report generated")
print("Excel:", OUTPUT_EXCEL)
print("Summary JSON:", SUMMARY_JSON)

